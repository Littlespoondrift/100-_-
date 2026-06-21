import jieba
import re
from openpyxl import load_workbook

# =========================
# 1. 加载自定义词典
# =========================
jieba.load_userdict("分词词典.txt")

# =========================
# 2. 加载停用词
# =========================
stopwords = set()
with open("中文停用词.txt", "r", encoding="utf-8") as f:
    for line in f:
        stopwords.add(line.strip())

# =========================
# 3. 读取CNKI Excel摘要
# =========================
file_path = "../智能交通_摘要数据.xlsx"

wb = load_workbook(file_path)
ws = wb.active

abstract_texts = []

for row in ws.iter_rows(min_row=2, values_only=True):
    title, abstract = row
    if abstract:
        abstract_texts.append(str(abstract))

print(f"读取摘要数量：{len(abstract_texts)}")

# =========================
# 4. 合并文本
# =========================
text = " ".join(abstract_texts)

# =========================
# 5. 清洗 + 分词
# =========================
def clean_and_tokenize(text):
    # 去符号
    text = re.sub(r"[^\w\s]", "", text)
    # 分词
    tokens = jieba.lcut(text)
    return tokens

# =========================
# 6. 去停用词
# =========================
def remove_stopwords(tokens):
    return [t for t in tokens if t not in stopwords and t.strip() != ""]

tokens = clean_and_tokenize(text)
tokens = remove_stopwords(tokens)

# =========================
# 7. 保存结果
# =========================
output_file = "清洗_智能交通文献摘要.txt"

with open(output_file, "w", encoding="utf-8") as f:
    f.write(" ".join(tokens))

print(f"清洗完成，已保存：{output_file}")